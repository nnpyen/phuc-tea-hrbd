#!/usr/bin/env python3
"""Extract HR + BD 2026 planning workbook into web-friendly JSON."""

from __future__ import annotations

import json
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

import openpyxl

SOURCE_XLSX = Path("/Users/nguyenngocphiyen/Downloads/KẾ HOẠCH 2026.xlsx")
OUTPUT_JSON = Path(
    "/Users/nguyenngocphiyen/Documents/Playground/hrbd-dashboard-2026-web/data/plan_2026.json"
)
OUTPUT_JS = Path(
    "/Users/nguyenngocphiyen/Documents/Playground/hrbd-dashboard-2026-web/data/plan_2026_data.js"
)


def is_blank(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, str) and value.strip() == "":
        return True
    return False


def as_number(value: Any) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        txt = value.strip()
        if txt in {"#REF!", "#DIV/0!", "#N/A"}:
            return None
        txt = txt.replace(",", "")
        try:
            return float(txt)
        except ValueError:
            return None
    return None


def as_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def normalize_month(value: Any) -> str:
    txt = as_text(value)
    if not txt:
        return ""
    txt = txt.replace("Thg ", "").replace("Tháng ", "").strip()
    return txt


def parse_scorecard(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    code_row: int,
    start_row: int,
    score_col: int,
) -> Dict[str, Any]:
    """Parse monthly target/actual blocks dynamically by KPI code row."""
    kpis: List[Dict[str, Any]] = []
    col = 3
    while col <= ws.max_column:
        code = as_text(ws.cell(code_row, col).value)
        if not code:
            # Stop after passing KPI blocks.
            if col > 20:
                break
            col += 1
            continue
        kpis.append(
            {
                "code": code,
                "target_col": col,
                "actual_col": col + 1,
                "score_col": col + 2,
                "manual_col": col + 3,
            }
        )
        col += 4

    months: List[Dict[str, Any]] = []
    row = start_row
    while row <= ws.max_row:
        month = normalize_month(ws.cell(row, 1).value)
        if not month:
            if row > start_row + 12:
                break
            row += 1
            continue

        month_item: Dict[str, Any] = {
            "month": month,
            "quarter": as_text(ws.cell(row, 2).value),
            "pillarScore": as_number(ws.cell(row, score_col).value),
            "kpis": {},
        }

        has_data = month_item["pillarScore"] is not None
        for item in kpis:
            target = as_number(ws.cell(row, item["target_col"]).value)
            actual = as_number(ws.cell(row, item["actual_col"]).value)
            score = as_number(ws.cell(row, item["score_col"]).value)
            manual = as_number(ws.cell(row, item["manual_col"]).value)
            if any(x is not None for x in (target, actual, score, manual)):
                has_data = True
            month_item["kpis"][item["code"]] = {
                "target": target,
                "actual": actual,
                "score": score,
                "manualScore": manual,
            }

        if has_data or row < start_row + 12:
            months.append(month_item)
        row += 1

    return {
        "kpiCodes": [k["code"] for k in kpis],
        "months": months,
    }


def main() -> None:
    wb = openpyxl.load_workbook(SOURCE_XLSX, data_only=True)

    setup = wb["SETUP"]
    year = int(as_number(setup["B3"].value) or 2026)
    weights = {
        "HR": as_number(setup["B5"].value) or 0.45,
        "VN": as_number(setup["B6"].value) or 0.45,
        "INTL": as_number(setup["B7"].value) or 0.10,
    }

    plan_ws = wb["PLAN_ANNUAL_QUARTER"]
    annual_plan: List[Dict[str, Any]] = []
    for row in range(7, 22):
        code = as_text(plan_ws.cell(row, 2).value)
        if not code:
            continue
        annual_plan.append(
            {
                "stt": as_text(plan_ws.cell(row, 1).value),
                "kpiCode": code,
                "pillar": as_text(plan_ws.cell(row, 3).value),
                "objective": as_text(plan_ws.cell(row, 4).value),
                "kpiName": as_text(plan_ws.cell(row, 5).value),
                "unit": as_text(plan_ws.cell(row, 6).value),
                "targetYear": as_number(plan_ws.cell(row, 7).value),
                "actualYear": as_number(plan_ws.cell(row, 8).value),
                "owner": as_text(plan_ws.cell(row, 10).value),
                "initiatives": as_text(plan_ws.cell(row, 11).value),
                "dependencies": as_text(plan_ws.cell(row, 12).value),
                "due": as_text(plan_ws.cell(row, 13).value),
                "status": as_text(plan_ws.cell(row, 14).value),
                "notes": as_text(plan_ws.cell(row, 15).value),
            }
        )

    quarter_plan: List[Dict[str, Any]] = []
    for row in range(25, 46):
        quarter_raw = as_text(plan_ws.cell(row, 1).value)
        pillar = as_text(plan_ws.cell(row, 3).value)
        if not quarter_raw or not pillar:
            continue
        quarter_plan.append(
            {
                "quarter": f"Q{quarter_raw}" if quarter_raw.isdigit() else quarter_raw,
                "pillar": pillar,
                "objective": as_text(plan_ws.cell(row, 4).value),
                "kpiName": as_text(plan_ws.cell(row, 5).value),
                "unit": as_text(plan_ws.cell(row, 6).value),
                "targetQuarter": as_number(plan_ws.cell(row, 7).value),
                "actualQuarter": as_number(plan_ws.cell(row, 8).value),
                "milestones": as_text(plan_ws.cell(row, 10).value),
                "keyActions": as_text(plan_ws.cell(row, 11).value),
                "owner": as_text(plan_ws.cell(row, 12).value),
                "due": as_text(plan_ws.cell(row, 13).value),
                "status": as_text(plan_ws.cell(row, 14).value),
                "notes": as_text(plan_ws.cell(row, 15).value),
            }
        )

    hr_scorecard = parse_scorecard(
        wb["HR_SCORECARD"], code_row=5, start_row=9, score_col=27
    )
    vn_scorecard = parse_scorecard(
        wb["VN_FRANCHISE_SCORECARD"], code_row=6, start_row=10, score_col=27
    )
    intl_scorecard = parse_scorecard(
        wb["INTERNATIONAL_SCORECARD"], code_row=6, start_row=10, score_col=15
    )

    pc_ws = wb["PARTNER_CARE_MONTHLY"]
    partner_plan: List[Dict[str, Any]] = []
    for row in range(7, 19):
        month = normalize_month(pc_ws.cell(row, 1).value)
        if not month:
            continue
        partner_plan.append(
            {
                "month": month,
                "quarter": as_text(pc_ws.cell(row, 2).value),
                "plannedNew": as_number(pc_ws.cell(row, 3).value),
                "actualNew": as_number(pc_ws.cell(row, 4).value),
                "complianceNew": as_number(pc_ws.cell(row, 5).value),
                "plannedExisting": as_number(pc_ws.cell(row, 6).value),
                "actualExisting": as_number(pc_ws.cell(row, 7).value),
                "complianceExisting": as_number(pc_ws.cell(row, 8).value),
                "notes": as_text(pc_ws.cell(row, 9).value),
            }
        )

    partner_logs: List[Dict[str, Any]] = []
    for row in range(23, 1000):
        month = normalize_month(pc_ws.cell(row, 2).value)
        partner_key = as_text(pc_ws.cell(row, 3).value)
        if not month and not partner_key:
            continue
        partner_type = as_text(pc_ws.cell(row, 4).value)
        if partner_type in {"#N/A", "#REF!"}:
            continue
        partner_logs.append(
            {
                "date": as_number(pc_ws.cell(row, 1).value),
                "month": month,
                "partnerKey": partner_key,
                "partnerType": partner_type,
                "partnerName": as_text(pc_ws.cell(row, 5).value),
                "touchpointType": as_text(pc_ws.cell(row, 6).value),
                "channel": as_text(pc_ws.cell(row, 7).value),
                "owner": as_text(pc_ws.cell(row, 8).value),
                "status": as_text(pc_ws.cell(row, 9).value),
                "dayDone": as_number(pc_ws.cell(row, 10).value),
                "result": as_text(pc_ws.cell(row, 11).value),
                "response": as_text(pc_ws.cell(row, 12).value),
            }
        )

    intl_ws = wb["INTL PROBLEM RECORD"]
    intl_problems: List[Dict[str, Any]] = []
    for row in range(2, intl_ws.max_row + 1):
        if all(is_blank(intl_ws.cell(row, col).value) for col in range(1, 6)):
            continue
        intl_problems.append(
            {
                "date": as_number(intl_ws.cell(row, 1).value),
                "month": normalize_month(intl_ws.cell(row, 2).value),
                "case": as_text(intl_ws.cell(row, 3).value),
                "result": as_text(intl_ws.cell(row, 4).value),
                "status": as_text(intl_ws.cell(row, 5).value),
                "category": "",
                "rootCause": "",
                "slaHours": None,
                "resolvedHours": None,
                "owner": "",
                "nextAction": "",
            }
        )

    summary_ws = wb["SUMMARY"]
    summary_by_month: List[Dict[str, Any]] = []
    for row in range(8, 20):
        month = normalize_month(summary_ws.cell(row, 1).value)
        if not month:
            continue
        summary_by_month.append(
            {
                "month": month,
                "quarter": as_text(summary_ws.cell(row, 2).value),
                "hrScore": as_number(summary_ws.cell(row, 3).value),
                "vnScore": as_number(summary_ws.cell(row, 4).value),
                "intlScore": as_number(summary_ws.cell(row, 5).value),
                "totalScore": as_number(summary_ws.cell(row, 6).value),
                "notes": as_text(summary_ws.cell(row, 7).value),
            }
        )

    data = {
        "meta": {
            "title": "HR + Business Development Dashboard 2026",
            "year": year,
            "updatedAt": datetime.now().isoformat(timespec="seconds"),
            "weights": weights,
            "sourceFile": str(SOURCE_XLSX),
        },
        "settings": {
            "thresholds": {
                "good": 0.9,
                "watch": 0.75,
            }
        },
        "annualPlan": annual_plan,
        "quarterPlan": quarter_plan,
        "scorecards": {
            "HR": hr_scorecard,
            "VN": vn_scorecard,
            "INTL": intl_scorecard,
        },
        "partnerCare": {
            "monthlyPlan": partner_plan,
            "logs": partner_logs,
        },
        "intlProblems": intl_problems,
        "summary": summary_by_month,
        "version": 1,
    }

    json_payload = json.dumps(data, ensure_ascii=False, indent=2)
    OUTPUT_JSON.write_text(json_payload, encoding="utf-8")
    OUTPUT_JS.write_text(
        "window.__PLAN_2026_DATA__ = " + json_payload + ";\n",
        encoding="utf-8",
    )
    print(f"Wrote: {OUTPUT_JSON}")
    print(f"Wrote: {OUTPUT_JS}")


if __name__ == "__main__":
    main()
